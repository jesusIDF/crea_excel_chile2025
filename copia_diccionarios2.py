import sys
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from copy import copy
import argparse
import re

# Definir los argumentos
parser = argparse.ArgumentParser(description="Copiar diccionarios a partir de IDF_Individual_Fields")
parser.add_argument("marca", help="Marca de auto a preservar")
parser.add_argument("archivo_entrada", help="Ruta del archivo Excel de entrada")
parser.add_argument("archivo_salida", help="Ruta del archivo Excel de salida")

# Parsear argumentos
args = parser.parse_args()

print(f"Marca a filtrar: {args.marca}")
print(f"Archivo de entrada: {args.archivo_entrada}")
print(f"Archivo de salida: {args.archivo_salida}")

# Archivos
#archivo_base = "../exceles/IDF_Individual_Fields.xlsx"
archivo_base = args.archivo_entrada
#archivo_destino = "../exceles/IDF_Chile_Chevrolet_2025.xlsx"
archivo_destino = args.archivo_salida
try:
    # Abrir el archivo base y obtener la hoja "Base Vehicle"
    wb_base = load_workbook(archivo_base)
    print("Cargando archivo base")
except FileNotFoundError:
    print(f"❌ Error: El archivo '{archivo_base}' no fue encontrado.")
    sys.exit(1)
except PermissionError:
    print(f"❌ Error: No tienes permisos para acceder a '{archivo_base}'.")
    sys.exit(1)
except InvalidFileException:
    print(f"❌ Error: El archivo '{archivo_base}' no es un archivo Excel válido o está dañado.")
    sys.exit(1)
except Exception as e:
    print(f"❌ Error inesperado al abrir '{archivo_base}': {e}")
    sys.exit(1)
    
# copias fieles
hoja_make_base : Worksheet = wb_base["Make"]
hoja_body_base : Worksheet = wb_base["Body"]
hoja_drive_base : Worksheet = wb_base["Drive"]
# filtrados
hoja_model_base: Worksheet = wb_base["Model"]
hoja_base: Worksheet = wb_base["Base Vehicle"]
hoja_vtobody_base : Worksheet = wb_base["V to Body"]
hoja_vtodrive_base : Worksheet = wb_base["V to Drive"]

try: 
    # Abrir el archivo destino
    wb_destino = load_workbook(archivo_destino)
    print("Cargando archivo destino")
except FileNotFoundError:
    print(f"❌ Error: El archivo '{archivo_destino}' no fue encontrado.")
    sys.exit(1)
except PermissionError:
    print(f"❌ Error: No tienes permisos para acceder a '{archivo_destino}'.")
    sys.exit(1)
except InvalidFileException:
    print(f"❌ Error: El archivo '{archivo_destino}' no es un archivo Excel válido o está dañado.")
    sys.exit(1)
except Exception as e:
    print(f"❌ Error inesperado al abrir '{archivo_destino}': {e}")
    sys.exit(1)

print("Copiando diccionarios ...")   
############# COPIAS TAL CUAL ###############
# Crear o reemplazar hoja destino Make
if "Make" in wb_destino.sheetnames:
    del wb_destino["Make"]
hoja_make_destino = wb_destino.create_sheet("Make")

# Copiar contenido celda por celda (incluyendo estilo)
for fila in hoja_make_base.iter_rows():
    for celda in fila:
        nueva_celda = hoja_make_destino.cell(row=celda.row, column=celda.column, value=celda.value)
        if celda.has_style:
            nueva_celda.font = copy(celda.font)
            nueva_celda.border = copy(celda.border)
            nueva_celda.fill = copy(celda.fill)
            nueva_celda.number_format = copy(celda.number_format)
            nueva_celda.protection = copy(celda.protection)
            nueva_celda.alignment = copy(celda.alignment)
print("✓ Hoja Make creada")

# Crear o reemplazar hoja destino Body
if "Body" in wb_destino.sheetnames:
    del wb_destino["Body"]
hoja_body_destino = wb_destino.create_sheet("Body")

# Copiar contenido celda por celda (incluyendo estilo)
for fila in hoja_body_base.iter_rows():
    for celda in fila:
        nueva_celda = hoja_body_destino.cell(row=celda.row, column=celda.column, value=celda.value)
        if celda.has_style:
            nueva_celda.font = copy(celda.font)
            nueva_celda.border = copy(celda.border)
            nueva_celda.fill = copy(celda.fill)
            nueva_celda.number_format = copy(celda.number_format)
            nueva_celda.protection = copy(celda.protection)
            nueva_celda.alignment = copy(celda.alignment)
print("✓ Hoja Body creada")

# Crear o reemplazar hoja destino Drive
if "Drive" in wb_destino.sheetnames:
    del wb_destino["Drive"]
hoja_drive_destino = wb_destino.create_sheet("Drive")

# Copiar contenido celda por celda (incluyendo estilo)
for fila in hoja_drive_base.iter_rows():
    for celda in fila:
        nueva_celda = hoja_drive_destino.cell(row=celda.row, column=celda.column, value=celda.value)
        if celda.has_style:
            nueva_celda.font = copy(celda.font)
            nueva_celda.border = copy(celda.border)
            nueva_celda.fill = copy(celda.fill)
            nueva_celda.number_format = copy(celda.number_format)
            nueva_celda.protection = copy(celda.protection)
            nueva_celda.alignment = copy(celda.alignment)
print("✓ Hoja Drive creada")

############# COPIAS FILTRADAS ###############
# Crear nueva hoja en el archivo destino
print("Filtrando Base Vehicle")
if "Vehi Base" in wb_destino.sheetnames:
    del wb_destino["Vehi Base"]  # Eliminar si ya existe para sobrescribir
hoja_nueva = wb_destino.create_sheet(title="Vehi Base")

filtro_marca = args.marca.strip().lower()
filtro_vtn = {"car", "truck", "van"}

# Copiar encabezado
for col in range(1, hoja_base.max_column + 1):
    valor = hoja_base.cell(row=1, column=col).value
    hoja_nueva.cell(row=1, column=col, value=valor)

# Filtrar y copiar filas
fila_destino = 2
for fila in range(2, hoja_base.max_row + 1):
    marca = hoja_base[f"G{fila}"].value
    tipo = hoja_base[f"J{fila}"].value

    if (
        isinstance(marca, str) and marca.strip().lower() == filtro_marca and
        isinstance(tipo, str) and tipo.strip().lower() in filtro_vtn
    ):
        for col in range(1, hoja_base.max_column + 1):
            val = hoja_base.cell(row=fila, column=col).value
            hoja_nueva.cell(row=fila_destino, column=col, value=val)
        fila_destino += 1
print("✓ Hoja Vehi Base creada")

# Obtener valores únicos de columna I de "Base Vehicle"
print("Extrayendo valores únicos de la columna I en 'Vehi Base'...")

hoja_model: Worksheet = wb_destino["Vehi Base"]
vehi_type_id = set()

for fila in range(2, hoja_model.max_row + 1):
    val = hoja_model[f"I{fila}"].value
    if val is not None:
        vehi_type_id.add(val)
print(f"✓ {len(vehi_type_id)} valores únicos encontrados {vehi_type_id}.")

if "Model" in wb_destino.sheetnames:
    del wb_destino["Model"]
hoja_model_destino = wb_destino.create_sheet("Model")

# Copiar encabezado
for col in range(1, hoja_model_base.max_column + 1):
    valor = hoja_model_base.cell(row=1, column=col).value
    hoja_model_destino.cell(row=1, column=col, value=valor)
    
# Filtrar y copiar filas según columna C
fila_destino = 2
for fila in range(2, hoja_model_base.max_row + 1):
    valor_C = hoja_model_base[f"C{fila}"].value
    if valor_C in vehi_type_id:
        for col in range(1, hoja_model_base.max_column + 1):
            hoja_model_destino.cell(
                row=fila_destino,
                column=col,
                value=hoja_model_base.cell(row=fila, column=col).value
            )
        fila_destino += 1
print("✓ Hoja Model creada y filtrada")

# V to Body
if "V to Body" in wb_destino.sheetnames:
    del wb_destino["V to Body"]
hoja_vtobody_destino = wb_destino.create_sheet("V to Body")

# Copiar encabezado
for col in range(1, hoja_vtobody_base.max_column + 1):
    valor = hoja_vtobody_base.cell(row=1, column=col).value
    hoja_vtobody_destino.cell(row=1, column=col, value=valor)

print(f"Filtrando 'V to Body con {filtro_marca} y {filtro_vtn}")
# Filtrar y copiar filas
fila_destino = 2
for fila in range(2, hoja_vtobody_base.max_row + 1):
    marca = hoja_vtobody_base[f"H{fila}"].value
    tipo = hoja_vtobody_base[f"K{fila}"].value

    if (
        isinstance(marca, str) and marca.strip().lower() == filtro_marca and
        isinstance(tipo, str) and tipo.strip().lower() in filtro_vtn
    ):
        for col in range(1, hoja_base.max_column + 1):
            val = hoja_vtobody_base.cell(row=fila, column=col).value
            hoja_vtobody_destino.cell(row=fila_destino, column=col, value=val)
        fila_destino += 1
print("✓ Hoja V to Body creada")

# V to Drive
if "V to Drive" in wb_destino.sheetnames:
    del wb_destino["V to Drive"]
hoja_vtodrive_destino = wb_destino.create_sheet("V to Drive")

# Copiar encabezado
for col in range(1, hoja_vtodrive_base.max_column + 1):
    valor = hoja_vtodrive_base.cell(row=1, column=col).value
    hoja_vtodrive_destino.cell(row=1, column=col, value=valor)

print(f"Filtrando 'V to Drive con {filtro_marca} y {filtro_vtn}")
# Filtrar y copiar filas
fila_destino = 2
for fila in range(2, hoja_vtodrive_base.max_row + 1):
    marca = hoja_vtodrive_base[f"G{fila}"].value
    tipo = hoja_vtodrive_base[f"J{fila}"].value

    if (
        isinstance(marca, str) and marca.strip().lower() == filtro_marca and
        isinstance(tipo, str) and tipo.strip().lower() in filtro_vtn
    ):
        for col in range(1, hoja_base.max_column + 1):
            val = hoja_vtodrive_base.cell(row=fila, column=col).value
            hoja_vtodrive_destino.cell(row=fila_destino, column=col, value=val)
        fila_destino += 1
print("✓ Hoja V to Drive creada")

# Guardar el archivo destino con la nueva hoja
wb_destino.save(archivo_destino)
exit()