from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from copy import copy
import argparse
import re
import sys

# Definir los argumentos
parser = argparse.ArgumentParser(description="Rasurar por columna Maker de un Excel")
parser.add_argument("marca", help="Marca de auto a preservar")
parser.add_argument("archivo_entrada", help="Ruta del archivo Excel de entrada")
parser.add_argument("archivo_salida", help="Ruta del archivo Excel de salida")

# Parsear argumentos
args = parser.parse_args()

print(f"Marca a filtrar: {args.marca}")
print(f"Archivo de entrada: {args.archivo_entrada}")
print(f"Archivo de salida: {args.archivo_salida}")

archivo = args.archivo_entrada
salida = args.archivo_salida
filtro = args.marca

# Cargar el archivo Excel
#archivo = "exceles/IDF_Chile_2025.xlsx"
try:
    wb = load_workbook(archivo)

    print("Estandarizando columna E")
except FileNotFoundError:
    print(f"❌ Error: El archivo '{archivo}' no fue encontrado.")
    sys.exit(1)
except PermissionError:
    print(f"❌ Error: No tienes permisos para acceder a '{archivo}'.")
    sys.exit(1)
except InvalidFileException:
    print(f"❌ Error: El archivo '{archivo}' no es un archivo Excel válido o está dañado.")
    sys.exit(1)
except Exception as e:
    print(f"❌ Error inesperado al abrir '{archivo}': {e}")
    sys.exit(1)

hoja = wb.active  # Usa la hoja activa    
# Color verde claro para relleno
relleno_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Recorre las celdas de la columna E (desde la segunda fila para omitir encabezado si existe)
for fila in range(2, hoja.max_row + 1):
    celda_E = hoja[f"E{fila}"]
    celda_C = hoja[f"C{fila}"]
    celda_K = hoja[f"K{fila}"]
    celda_L = hoja[f"L{fila}"]
    if celda_E.value and isinstance(celda_E.value, str):
        celda_E.value = celda_E.value.upper()
    # Si columna C (columna 3) tiene "IDF - Loaded", aplicar fondo verde
    valor_C = hoja.cell(row=fila, column=3).value
    #if valor_C == "IDF - Loaded":
    if celda_C.value and celda_C.value == "IDF - Loaded":
        celda_C.fill = relleno_verde
    celda_K.value = "CHL"
    celda_L.value = "Chile"

# Guardar los cambios en un nuevo archivo (opcional)
wb.save(archivo)

# filtrando
# Crear nuevo libro para guardar resultados
nuevo_wb = Workbook()
nueva_hoja = nuevo_wb.active
nueva_hoja.title = "IDF_"+ filtro + "_2025"

# Copiar encabezado
for col in range(1, hoja.max_column + 1):
    celda_origen = hoja.cell(row=1, column=col)
    celda_destino = nueva_hoja.cell(row=1, column=col, value=celda_origen.value)
    
    #nueva_hoja.cell(row=1, column=col, value=hoja.cell(row=1, column=col).value)
    # Copiar estilo
    if celda_origen.has_style:
        celda_destino.font = copy(celda_origen.font)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.alignment = copy(celda_origen.alignment)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.protection = copy(celda_origen.protection)

print("Estandarizada columna E")
# Compilar la expresión regular (ignorando mayúsculas/minúsculas)
patron = re.compile(args.marca, re.IGNORECASE)
print(f"Filtrando por {filtro}")
# Filtrar filas que coincidan con la expresión regular en columna E
nueva_fila = 2
for fila in range(2, hoja.max_row + 1):
    celda_E = hoja[f"E{fila}"]
    valor_E = celda_E.value
    if valor_E and isinstance(valor_E, str) and patron.search(valor_E):
        for col in range(1, hoja.max_column + 1):
            val = hoja.cell(row=fila, column=col).value
            # Convertir columna E a mayúsculas
            if col == 5 and isinstance(val, str):
                val = val.upper()
            #nueva_hoja.cell(row=nueva_fila, column=col, value=val)
            celda_nueva = nueva_hoja.cell(row=nueva_fila, column=col, value=val)

            # Si columna C (columna 3) tiene "IDF - Loaded", aplicar fondo verde
            valor_C = hoja.cell(row=fila, column=3).value
            if valor_C == "IDF - Loaded":
                celda_nueva.fill = relleno_verde
        nueva_fila += 1
print(f"Filtrado por {filtro}")

# Guardar los cambios en un nuevo archivo (opcional)
nuevo_wb.save(salida)
exit()