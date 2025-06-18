import sys
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
import argparse
from copy import copy
import re

# Definir los argumentos
parser = argparse.ArgumentParser(description="Copiar diccionarios a partir de IDF_Individual_Fields")
parser.add_argument("marca", help="Marca de auto a preservar")
parser.add_argument("archivo_entrada", help="Ruta del archivo Excel de entrada")
#parser.add_argument("archivo_salida", help="Ruta del archivo Excel de salida")

# Parsear argumentos
args = parser.parse_args()

print(f"Marca a filtrar: {args.marca}")
print(f"Archivo de entrada: {args.archivo_entrada}")
#print(f"Archivo de salida: {args.archivo_salida}")

archivo_base = args.archivo_entrada
try:
    wb_base = load_workbook(archivo_base)
    print("Cargando excel base")
except FileNotFoundError:
    print(f"❌ Error: El archivo '{archivo_base}' no fue encontrado.")
    sys.exit(1)
except PermissionError:
    print(f"❌ Error: No tienes permisos para acceder a '{archivo_base}'.")
    sys.exit(1)
except InvalidFileException:
    print(f"❌ Error: El archivo '{archivo_base}' no es un archivo Excel válido o está dañado.")
    sys.exit(1)
except Exception as e:
    print(f"❌ Error inesperado al abrir '{archivo_base}': {e}")
    sys.exit(1)
    
# Crear nuevo libro para guardar resultados
#wb_base = Workbook()
nueva_hoja = wb_base.active

# crear nuevas columnas despues de Make Name
# Insertar después de la columna E (índice 5) → insertamos en 6
nueva_hoja.insert_cols(6)
nueva_hoja.cell(row=1, column=6, value="Marca ACES")  # Encabezado para nueva columna después de E
nueva_hoja.insert_cols(7)
nueva_hoja.cell(row=1, column=7, value="VLOOKUP Make")  # Encabezado para nueva columna después de E
# =SI.ERROR(BUSCARV(F2,Make!$B:$B,1,0), -1)
nueva_hoja.cell(row=2, column=6, value='=ESPACIOS(E3)')
nueva_hoja.cell(row=2, column=7, value='=SI.ERROR(BUSCARV(F3,Make!B:B,1,0), -1)')
print("✓ Columnas ACES y vlookup para Make creadas")

# Crear nuevas columnas despues de Model Name
# Insertar después de la columna F original (ahora está en índice 8) → insertamos en 9
nueva_hoja.insert_cols(9)
nueva_hoja.cell(row=1, column=9, value="Modelo ACES")  # Encabezado para nueva columna después de H
nueva_hoja.insert_cols(10)
nueva_hoja.cell(row=1, column=10, value="VLOOKUP Model")  # Encabezado para nueva columna después de H
# =SI.ERROR(BUSCARV(I2,Model!$B:$B,1,0), -1)
nueva_hoja.cell(row=2, column=9, value='=ESPACIOS(H3)')
nueva_hoja.cell(row=2, column=10, value='=SI.ERROR(BUSCARV(I3,Model!B:B,1,0), -1)')
print("✓ Columnas ACES y vlookup para Model creadas")

# Crear nuevas columnas despues de V TypeName
# Insertar después de la columna G original (ahora está en índice 11) → insertamos en 12
nueva_hoja.insert_cols(12)
nueva_hoja.cell(row=1, column=12, value="V TypeName ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(13)
nueva_hoja.cell(row=1, column=13, value="VLOOKUP V TypeName")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(L2,'Vehi Base'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('Vehi Base'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=13, value='=SI.ERROR(BUSCARV(L3,\'Vehi Base\'!J:J,1,0), -1)')
print("✓ Columnas ACES y vlookup para V TypeName creadas")

# Crear nuevas columnas despues de V TypeGroup
# Insertar después de la columna H original (ahora está en índice 14) → insertamos en 15
nueva_hoja.insert_cols(15)
nueva_hoja.cell(row=1, column=15, value="V TypeGroup ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(16)
nueva_hoja.cell(row=1, column=16, value="VLOOKUP V TypeGroup")  # Encabezado para nueva columna después de F
#nueva_hoja.cell(row=2, column=16, value='=SI.ERROR(BUSCARV(O3,\'Vehicle Type Grp\'!B:B,1,0), -1)')
# no se revisa
print("✓ Columnas ACES y vlookup para V TypeGroup creadas")

# Crear nuevas columnas despues de SubmodelName
# Insertar después de la columna I original (ahora está en índice 17) → insertamos en 18
nueva_hoja.insert_cols(18)
nueva_hoja.cell(row=1, column=18, value="SubModelName ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(19)
nueva_hoja.cell(row=1, column=19, value="VLOOKUP SubModelName")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=19, value='=SI.ERROR(BUSCARV(R3,Submodel!B:B,1,0), -1)')
print("✓ Columnas ACES y vlookup para SubModelName creadas")

# Crear nuevas columnas despues de BodyNumDoors
# Insertar después de la columna I original (ahora está en índice 26) → insertamos en 27
nueva_hoja.insert_cols(27)
nueva_hoja.cell(row=1, column=27, value="BodyNumDoors ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(28)
nueva_hoja.cell(row=1, column=28, value="VLOOKUP BodyNumDoors")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=28, value='=SI.ERROR(BUSCARV(TEXTO(AA3,"0"),\'V to Body\'!E:E,1,0), -1)')
print("✓ Columnas ACES y vlookup para BodyNumDoors creadas")

# Crear nuevas columnas despues de BodyTypeName
# Insertar después de la columna I original (ahora está en índice 29) → insertamos en 30
nueva_hoja.insert_cols(30)
nueva_hoja.cell(row=1, column=30, value="BodyTypeName ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(31)
nueva_hoja.cell(row=1, column=31, value="VLOOKUP BodyTypeName")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=31, value='=SI.ERROR(BUSCARV(AD3,Body!B:B,1,0), -1)')
print("✓ Columnas ACES y vlookup para BodyTypeName creadas")

# Crear nuevas columnas despues de DriveTypeName
# Insertar después de la columna I original (ahora está en índice 32) → insertamos en 33
nueva_hoja.insert_cols(33)
nueva_hoja.cell(row=1, column=33, value="DriveTypeName ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(34)
nueva_hoja.cell(row=1, column=34, value="VLOOKUP DriveTypeName")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=34, value='=SI.ERROR(BUSCARV(AG3,Drive!B:B,1,0), -1)')
print("✓ Columnas ACES y vlookup para DriveTypeName creadas")

# Crear nuevas columnas despues de Liter
# Insertar después de la columna I original (ahora está en índice 46) → insertamos en 47
nueva_hoja.insert_cols(47)
nueva_hoja.cell(row=1, column=47, value="Liter ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(48)
nueva_hoja.cell(row=1, column=48, value="VLOOKUP Liter")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=48, value='=SI.ERROR(BUSCARV(TEXTO(AU3,"0.0"),\'V To Engine\'!H:H,1,0), -1)')
print("✓ Columnas ACES y vlookup para Liter creadas")

# Crear nuevas columnas despues de FuelTypeName
# Insertar después de la columna I original (ahora está en índice 68) → insertamos en 69
nueva_hoja.insert_cols(69)
nueva_hoja.cell(row=1, column=69, value="FuelTypeName ACES")  # Encabezado para nueva columna después de F
nueva_hoja.cell(row=2, column=69, value='=ESPACIOS(MAYUSC(BP3))')
nueva_hoja.insert_cols(70)
nueva_hoja.cell(row=1, column=70, value="VLOOKUP FuelTypeName")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=70, value='=SI.ERROR(BUSCARV(BQ3,Fuel!B:B,1,0), -1)')
print("✓ Columnas ACES y vlookup para FuelTypeName creadas")

# Crear nuevas columnas despues de TransmissionNumSpeeds
# Insertar después de la columna I original (ahora está en índice 79) → insertamos en 80
nueva_hoja.insert_cols(80)
nueva_hoja.cell(row=1, column=80, value="TransmissionNumSpeeds ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(81)
nueva_hoja.cell(row=1, column=81, value="VLOOKUP TransmissionNumSpeeds")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=81, value='=SI.ERROR(BUSCARV(TEXTO(CB3,"0"),\'V To Transmission\'!E:E,1,0), -1)')
print("✓ Columnas ACES y vlookup para TransmissionNumSpeeds creadas")

# Crear nuevas columnas despues de TransmissionControlTypeName
# Insertar después de la columna I original (ahora está en índice 82) → insertamos en 83
nueva_hoja.insert_cols(83)
nueva_hoja.cell(row=1, column=83, value="TransmissionControlTypeName ACES")  # Encabezado para nueva columna después de F
nueva_hoja.insert_cols(84)
nueva_hoja.cell(row=1, column=84, value="VLOOKUP TransmissionControlTypeName")  # Encabezado para nueva columna después de F
# =SI.ERROR(BUSCARV(R2,'V to Body'!$J:$J,1,0), -1)
# =SI(ESNUMERO(COINCIDIR(R2,UNICOS('V to Body'!$J:$J),0)), R2, -1)
nueva_hoja.cell(row=2, column=84, value='=SI.ERROR(BUSCARV(CE3,\'V To Transmission\'!D:D,1,0), -1)')
print("✓ Columnas ACES y vlookup para TransmissionControlTypeName creadas")

# IDF comments se deja, originalmente columna J (ahora esta en el indice 20)
# RegionAbbr se deja, originalmente columna K (ahora esta en el indice 21)
# RegionName se deja, originalmente columna L (ahora esta en el indice 22)
# Copiar la primera linea de IDF_Chile_2025
# Cargar libros
#wb_base = load_workbook(archivo_base)
wb_base_chile = load_workbook('../exceles/IDF_Chile_Chevrolet_2025_06_11.xlsx')

# Seleccionar la primera hoja de cada libro
hoja_destino = wb_base.active
hoja_origen = wb_base_chile.active
hoja_destino.insert_rows(1)  # Inserta una fila al inicio

print("Poniendo primer renglón final...")

# Copiar el contenido y estilo de la fila 1 (renglón 1)
for col in range(1, hoja_origen.max_column + 1):
    celda_origen = hoja_origen.cell(row=1, column=col)
    celda_destino = hoja_destino.cell(row=1, column=col, value=celda_origen.value)

    # Copiar estilos si los hay
    if celda_origen.has_style:
        celda_destino.font = copy(celda_origen.font)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.alignment = copy(celda_origen.alignment)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.protection = copy(celda_origen.protection)

print(f"Guardando archivo final {archivo_base}")
# Guardar el archivo destino con la nueva hoja
wb_base.save(archivo_base)
exit()