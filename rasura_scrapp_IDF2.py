from openpyxl import load_workbook
import pandas as pd
import argparse
import re
import os
import sys

# rasura_scrapp_IDF2.py Ford ../exceles/IDF_Chile_2025.xlsx ../exceles/IDF_Chile_Ford_2025.xlsx        
# Definir los argumentos
parser = argparse.ArgumentParser(description="Rasurar por columna Maker de un Excel")
parser.add_argument("marca", help="Marca de auto a preservar")
parser.add_argument("archivo_entrada", help="Ruta del archivo Excel de entrada")
parser.add_argument("archivo_salida", help="Ruta del archivo Excel de salida")

# Parsear argumentos
args = parser.parse_args()

print(f"Marca a filtrar: {args.marca}")
print(f"Archivo de entrada: {args.archivo_entrada}")
print(f"Archivo de salida: {args.archivo_salida}")

archivo = args.archivo_entrada
salida = args.archivo_salida
filtro = args.marca
# Extraer el nombre sin la extensión
nombre_archivo = os.path.splitext(archivo)[0]
nombre_salida = os.path.splitext(salida)[0]

# paso 1 convertir a CSV
try:
    # Esto es pesado, solo para conversión inicial
    df = pd.read_excel(archivo, engine='openpyxl')
    df.to_csv(nombre_archivo + '.csv', index=False)
    print(f"Creando archivo {nombre_archivo + '.csv'}")
except pd.errors.EmptyDataError:
    print("⚠️ El archivo está vacío.")
    sys.exit(1)
except pd.errors.ParserError as e:
    print(f"❌ Error de formato al parsear el CSV: {e}")
    sys.exit(1)
except UnicodeDecodeError:
    print("❌ Error de codificación: prueba con encoding='utf-8' o 'latin1'")
    sys.exit(1)
except Exception as e:
    print(f"❌ Ocurrió un error inesperado: {e}")
    sys.exit(1)
    
# paso 2 filtrar chuncks de datos con pandas
def filtrar_csv_por_subcadena(ruta_csv, subcadena, col='MakeName'):
    resultado = []
    for chunk in pd.read_csv(ruta_csv, chunksize=10000):
        filtrado = chunk[chunk[col].astype(str).str.contains(subcadena, case=False, na=False)]
        resultado.append(filtrado)
    
    # Concatenar todos los resultados filtrados
    df_resultado = pd.concat(resultado, ignore_index=True)
    return df_resultado

print(f"Filtrando archivo {nombre_archivo + '.csv'}")
df_filtrado = filtrar_csv_por_subcadena(nombre_archivo + '.csv', filtro.strip().lower())

print(f"Filtrando columna E por {filtro.strip().lower()}")

# Modificar columnas E, K y L
df_filtrado['MakeName'] = df_filtrado['MakeName'].str.upper()
df_filtrado['KRegionAbbr'] = 'CHL'
df_filtrado['RegionName'] = 'Chile'

# agregamos columnas
def agrega_columnas(df, columna_referencia, nuevas_columnas, valor_vacio=''):
    """
    Inserta columnas vacías después de una columna existente en un DataFrame.
    
    Parámetros:
        df (pd.DataFrame): DataFrame original
        columna_referencia (str): Nombre de la columna después de la cual se insertarán las nuevas
        nuevas_columnas (list): Lista de nombres de columnas nuevas a insertar
        valor_vacio (any): Valor que se usará para llenar las columnas nuevas (por defecto: '')

    Retorna:
        pd.DataFrame: El DataFrame modificado
    """
    df = df.copy()
    
    if columna_referencia not in df.columns:
        raise ValueError(f"La columna '{columna_referencia}' no existe en el DataFrame.")
    
    pos = df.columns.get_loc(columna_referencia) + 1
    
    for i, nombre_col in enumerate(nuevas_columnas):
        df.insert(loc=pos + i, column=nombre_col, value=[valor_vacio] * len(df))

    return df

df_modificado = agrega_columnas(df_filtrado, 'MakeName', ['Marca ACES', 'VLOOKUP Marca'])
#num_columnas = agrega_columnas('Make', ['Marca ACES', 'VLOOKUP Marca'])
df_modificado = agrega_columnas(df_modificado, 'ModelName', ['Modelo ACES', 'VLOOKUP Modelo'])
df_modificado = agrega_columnas(df_modificado, 'VehicleTypeName', ['VehicleType ACES', 'VLOOKUP VehicleType'])
df_modificado = agrega_columnas(df_modificado, 'SubmodelName', ['SubmodelName ACES', 'VLOOKUP SubmodelName'])
print(f"Nuevo CSV con {df_modificado.shape[1]} columnas")

df_modificado.to_csv(nombre_salida + '.csv', index=False)
print(f"✅ Resultado guardado en {nombre_salida + '.csv'} con {len(df_modificado)} filas.")
print("Guardando archivo CSV")
# Guardar como archivo Excel (.xlsx)
#df_modificado.to_excel(salida, index=False, engine='openpyxl')
with pd.ExcelWriter(salida, engine='openpyxl') as writer:
    df_modificado.to_excel(writer, sheet_name=nombre_salida, index=False)
print("Guardando de CSV a Excel")
# copiar el encabezado del base
from copy import copy

def copiar_encabezado(hoja_origen, hoja_destino):
    """
    Copia la primera fila (encabezado) de hoja_origen a hoja_destino,
    incluyendo el valor y los estilos de cada celda.
    """
    for fila in hoja_origen.iter_rows(min_row=1, max_row=1):
        for celda in fila:
            nueva_celda = hoja_destino.cell(row=1, column=celda.column, value=celda.value)
            if celda.has_style:
                nueva_celda.font = copy(celda.font)
                nueva_celda.border = copy(celda.border)
                nueva_celda.fill = copy(celda.fill)
                nueva_celda.number_format = copy(celda.number_format)
                nueva_celda.protection = copy(celda.protection)
                nueva_celda.alignment = copy(celda.alignment)

wb = load_workbook(archivo)
#hoja = wb[nombre_salida]
hoja = wb.active
nueva_hoja = wb.create_sheet('EncabezadoCopiado')

copiar_encabezado(hoja, nueva_hoja)

wb.save('archivo_modificado.xlsx')
exit()